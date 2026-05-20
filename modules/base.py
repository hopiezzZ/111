# modules/base.py
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QLabel, QTableWidget, QPushButton, QFileDialog, QMessageBox, QProgressDialog
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import tempfile
import os
import shutil

# ========== 报表样式函数 ==========
def apply_excel_style(workbook):
    """对工作簿应用样式：标题加粗居中、边框、合并PDT二级分类列、自动列宽"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        # 1. 标题行样式
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        # 2. 数据区域边框和对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 3. 针对 PDT产品线 合并二级分类列（第一列）+ 订单目标(21) + 完成率(22)
        if sheet_name == 'PDT产品线' and ws.max_row > 1:
            col_idx = 1
            start_row = 2
            current_val = ws.cell(row=start_row, column=col_idx).value
            for row in range(start_row+1, ws.max_row+2):
                val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
                if val != current_val:
                    if row - 1 > start_row:
                        ws.merge_cells(start_row=start_row, start_column=col_idx,
                                       end_row=row-1, end_column=col_idx)
                        ws.merge_cells(start_row=start_row, start_column=21,
                                       end_row=row-1, end_column=21)
                        ws.merge_cells(start_row=start_row, start_column=22,
                                       end_row=row-1, end_column=22)
                        for r in range(start_row, row):
                            ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=r, column=21).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=r, column=22).alignment = Alignment(horizontal='center', vertical='center')
                    start_row = row
                    current_val = val

        # 3.5. PDT 服务产品经理业绩总计行合并它和右边单元格
        if sheet_name == 'PDT产品线' and ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == '服务产品经理业绩总计':
                    ws.merge_cells(start_row=row, start_column=1,
                                   end_row=row, end_column=2)
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    break

        # 4. 针对 行业计算表 / 行业报表 合并事业部列（第一列）
        if sheet_name in ('行业计算表', '行业报表') and ws.max_row > 1:
            col_idx = 1
            start_row = 2
            current_val = ws.cell(row=start_row, column=col_idx).value
            for row in range(start_row+1, ws.max_row+2):
                val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
                if val != current_val:
                    if row - 1 > start_row:
                        ws.merge_cells(start_row=start_row, start_column=col_idx,
                                       end_row=row-1, end_column=col_idx)
                        for r in range(start_row, row):
                            ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                    start_row = row
                    current_val = val

        # 5. 针对 办事处计算表 / 办事处报表 合并区域列（第一列）
        if sheet_name in ('办事处计算表', '办事处报表') and ws.max_row > 1:
            col_idx = 1
            start_row = 2
            current_val = ws.cell(row=start_row, column=col_idx).value
            for row in range(start_row+1, ws.max_row+2):
                val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
                if val != current_val:
                    if row - 1 > start_row:
                        ws.merge_cells(start_row=start_row, start_column=col_idx,
                                       end_row=row-1, end_column=col_idx)
                        for r in range(start_row, row):
                            ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                    start_row = row
                    current_val = val

        # 6. 自动列宽（限制最大宽度30）
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


# ========== 通用报表生成线程 ==========
class ReportThread(QThread):
    finished = pyqtSignal(object)
    error = pyqtSignal(str)

    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class BaseModule(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_df = None
        self.current_report_path = None
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f5f5;
                color: #000000;
            }
            QLabel {
                color: #000000;
            }
            QTableWidget {
                background-color: #ffffff;
                alternate-background-color: #f9f9f9;
                gridline-color: #ddd;
            }
            QTableWidget::item {
                color: #000000;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                color: #000000;
                border: 1px solid #ccc;
            }
            QPushButton {
                background-color: #e0e0e0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 6px 12px;
                color: #000000;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QProgressDialog {
                background-color: #ffffff;
                color: #000000;
            }
        """)
        self.init_ui()
        self.setAcceptDrops(True)

    def init_ui(self):
        pass

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                self.load_file(file_path)
            else:
                QMessageBox.warning(self, "格式错误", "请上传 .xlsx 或 .xls 文件")

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", os.path.expanduser("~"), "Excel文件 (*.xlsx *.xls)",
            options=QFileDialog.DontUseNativeDialog
        )
        if file_path:
            self.load_file(file_path)

    def load_file(self, file_path):
        try:
            df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
            self.current_df = df
            self.on_file_loaded(df)
        except Exception as e:
            QMessageBox.critical(self, "读取失败", str(e))

    def on_file_loaded(self, df):
        pass

    def save_report(self):
        if not self.current_report_path or not os.path.exists(self.current_report_path):
            QMessageBox.warning(self, "无报表", "请先生成报表")
            return
        desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
        default_path = os.path.join(desktop, "报表结果.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存报表", default_path, "Excel文件 (*.xlsx)",
            options=QFileDialog.DontUseNativeDialog
        )
        if save_path:
            shutil.copy2(self.current_report_path, save_path)
            QMessageBox.information(self, "保存成功", f"报表已保存至：{save_path}")